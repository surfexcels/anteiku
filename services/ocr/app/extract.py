from __future__ import annotations

import io
from typing import Literal

import pdfplumber
from openpyxl import load_workbook
from pdf2image import convert_from_bytes
import pytesseract

from app.line_parser import parse_csv, parse_pdf_tables, parse_text_lines
from app.models import ExtractionResponse, LineItem

Method = Literal["csv", "excel", "pdfplumber", "tesseract"]


def _extract_pdf(content: bytes) -> tuple[str, list[LineItem], Method]:
    text_parts: list[str] = []
    table_items: list[LineItem] = []

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if page_text:
                text_parts.append(page_text)
            tables = page.extract_tables() or []
            table_items.extend(parse_pdf_tables([tables]))

    text = "\n".join(text_parts).strip()
    if len(text) >= 80 or table_items:
        line_items = table_items or parse_text_lines(text)
        return text, line_items, "pdfplumber"

    ocr_chunks: list[str] = []
    for image in convert_from_bytes(content, dpi=200):
        ocr_chunks.append(pytesseract.image_to_string(image))

    text = "\n".join(ocr_chunks).strip()
    return text, parse_text_lines(text), "tesseract"


def _extract_excel(content: bytes) -> tuple[str, list[LineItem], Method]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return "", [], "excel"

    text_lines = [" | ".join(str(cell or "") for cell in row) for row in rows]
    text = "\n".join(text_lines)

    header = [str(cell or "").strip() for cell in rows[0]]
    table = [header] + [[str(cell or "").strip() for cell in row] for row in rows[1:]]
    line_items = parse_pdf_tables([table])
    return text, line_items or parse_text_lines(text), "excel"


def extract_document(filename: str, content: bytes) -> ExtractionResponse:
    lower = filename.lower()

    if lower.endswith(".csv"):
        text, line_items = parse_csv(content)
        return ExtractionResponse(method="csv", text=text, line_items=line_items)

    if lower.endswith((".xlsx", ".xlsm", ".xltx")):
        text, line_items, method = _extract_excel(content)
        return ExtractionResponse(method=method, text=text, line_items=line_items)

    if lower.endswith(".pdf"):
        text, line_items, method = _extract_pdf(content)
        return ExtractionResponse(method=method, text=text, line_items=line_items)

    raise ValueError(f"Unsupported file type: {filename}")
